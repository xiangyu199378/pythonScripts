"""
基金数据管理
@author:Mark Xiang 
"""
import  os
import  xlrd
import  requests
import re
import json
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib
from turtle import down
from bs4 import BeautifulSoup
from xlutils.copy import copy
import xlsxwriter
import openpyxl as op
from  sql import sql_table_fund
# 处理乱码
matplotlib.rcParams['font.sans-serif'] = ['SimHei']
matplotlib.rcParams['font.family']='sans-serif'
matplotlib.rcParams['axes.unicode_minus'] = False

FundList=[] 

def get_html(code):
    Url = 'http://fund.eastmoney.com/{0}.html'.format(code)
    rsp = requests.get(Url)
    rsp.encoding = "UTF-8" 
    html = rsp.text  
    return html



def deal_excel(path):
    # 打开基金原始表格
    rdBook=xlrd.open_workbook(path)
    rdSheet=rdBook.sheet_by_name('基金数据')
    dstb=copy(rdBook)
    dstt=dstb.get_sheet(2)

    #行row，列cols
    #print("nrows",rdSheet.nrows)
    #print("ncols",rdSheet.ncols)   
    curr_dir = os.getcwd()
    wb = op.load_workbook(path)
    sh=wb['基金数据']

    row_j = 2
    fundCnt=rdSheet.nrows 
    for j in range(row_j,fundCnt):
        FundCode=str(int(rdSheet.cell_value(j,1)))
        if len(FundCode)>0:
            code=('0'*(6-len(FundCode))+FundCode)
            html=get_html(code)
            soup = BeautifulSoup(html,'html.parser') 
            # 基金净值信息
            price_dl= soup.find("dl",class_="dataItem02")
            p_text = price_dl.find("p")
            price_spans=price_dl.find("dd",class_="dataNums").find_all("span")
            # 基金日期
            update_date = p_text.get_text().replace("单位净值 (","").replace(")","")
            sh.cell(j,8,update_date)
            # 基金净值
            price = price_spans[0].get_text()
            sh.cell(j,7,price)
            # 基金涨幅
            price_percent = price_spans[1].get_text().replace("%","")
            sh.cell(j,6,price_percent) 
        else:
            print("基金代码不得为空")
            exit(0)
  
    wb.save(path)
    return  rdBook

if __name__ == '__main__':
    curr_dir = os.getcwd()
   #book = deal_excel(curr_dir + '\向氏集团投资跟踪表.xlsx')
    sql_table_fund(1)

